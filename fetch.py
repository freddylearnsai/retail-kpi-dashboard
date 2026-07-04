"""UCI Online Retail II: download zip, stream both xlsx sheets into DuckDB → data/retail.parquet."""
import io, json, time, urllib.request, zipfile
import duckdb
from openpyxl import load_workbook

URL = "https://archive.ics.uci.edu/static/public/502/online+retail+ii.zip"
COLS = ["Invoice", "StockCode", "Description", "Quantity", "InvoiceDate", "Price", "Customer ID", "Country"]

def main() -> None:
    t0 = time.time()
    with urllib.request.urlopen(URL, timeout=300) as r:
        z = zipfile.ZipFile(io.BytesIO(r.read()))
    xlsx = [n for n in z.namelist() if n.endswith(".xlsx")][0]
    wb = load_workbook(io.BytesIO(z.read(xlsx)), read_only=True, data_only=True)
    con = duckdb.connect()
    con.execute("""create table t (invoice varchar, stock_code varchar, description varchar,
                   quantity bigint, invoiced_at timestamp, price double, customer_id varchar, country varchar)""")
    rows = 0
    for sheet in wb.worksheets:
        it = sheet.iter_rows(values_only=True)
        header = next(it)
        assert list(header) == COLS, f"unexpected header {header}"
        batch = []
        for row in it:
            batch.append(row)
            if len(batch) == 50000:
                con.executemany("insert into t values (?,?,?,?,?,?,?,?)", batch); rows += len(batch); batch = []
        if batch:
            con.executemany("insert into t values (?,?,?,?,?,?,?,?)", batch); rows += len(batch)
    con.execute("copy t to 'data/retail.parquet' (format parquet)")
    span = con.execute("select min(invoiced_at), max(invoiced_at) from t").fetchone()
    inv = con.execute("select count(distinct invoice) from t").fetchone()[0]
    print(json.dumps({"rows": rows, "invoices": inv, "span": [str(span[0]), str(span[1])],
                      "sheets": len(wb.worksheets), "seconds": round(time.time() - t0, 1)}))

if __name__ == "__main__":
    main()
